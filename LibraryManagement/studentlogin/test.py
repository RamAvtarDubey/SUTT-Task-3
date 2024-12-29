import openpyxl


book = openpyxl.load_workbook('./booklist.xlsx')
sheet = book.active
# cells = sheet['B2': 'E3']
books_details_list = []
for row in sheet.iter_rows(min_row=2, max_col=6, values_only=True):
    name = row[1]  # First column as 'name'
    isbn = row[2]
    author = row[3]
    publisher = row[4]
    number = row[5] 
    # print(name)
    # if name and any(detail):  # Only include rows with a name and at least one detail
    books_details_list.append({'name': name, 'isbn': isbn, 'author': author, 'publisher': publisher, 'number': number})
print(books_details_list)


# book = openpyxl.load_workbook('./booklist.xlsx')
# sheet = book.active
# # cells = sheet['B2': 'E3']
# books_details_list = []

# for row in sheet.iter_rows(min_row=2, max_col=5, values_only=True):
#     name = row[1]  # First column as 'name'
#     detail = list(row[2:5])  # Next four columns as 'detail'
#     # print(name)
#     # if name and any(detail):  # Only include rows with a name and at least one detail
#     books_details_list.append({'name': name, 'detail': detail})


# print(books_details_list)