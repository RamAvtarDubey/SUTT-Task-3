import os
import openpyxl
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.views.decorators.csrf import csrf_exempt
from google.oauth2 import id_token
from google.auth.transport import requests
import pandas as pd
from django.template.loader import render_to_string

def generate_html_files(details):
    base_dir = 'studentlogin/static/studentlogin/generated_pages'
    if not os.path.exists(base_dir):
        os.makedirs(base_dir)

    for i, detail in enumerate(details):
        context = {'detail': detail}
        html_content = render_to_string('studentlogin/detail_template.html', context)
        file_path = os.path.join(base_dir, f'detail_{i}.html')
        with open(file_path, 'w') as file:
            file.write(html_content)


def home(request):
    return render(request,'studentlogin/home.html')

def about(request):
    return render(request,'studentlogin/about.html')

def admin_login(request):
    try:
        book = openpyxl.load_workbook('./booklist.xlsx')
        sheet = book.active
        # cells = sheet['B2': 'E3']
        books_details_list = []
        for row in sheet.iter_rows(min_row=2, max_col=6, values_only=True):
            name = row[1]  # First column as 'name'
            isbn = row[2]
            author = row[3]
            publisher = row[4]
            number = row[5] 
            # print(name)
            # if name and any(detail):  # Only include rows with a name and at least one detail
            books_details_list.append({'name': name, 'isbn': isbn, 'author': author, 'publisher': publisher, 'number': number})


        

        # Generate HTML files for each detail
        generate_html_files(books_details_list)

        return render(request, 'studentlogin/admin_login.html', {'details': books_details_list})
    except Exception as e:
        return render(request, 'studentlogin/admin_login.html', {'error': str(e)})
    
def upload_excel(request):
    return render(request,'studentlogin/upload_excel.html')


def done_uploading_excel(request):

    if request.method == 'POST':
        file = request.FILES['excel_file']
        csv = pd.read_excel(file)
        names = csv['Name']
        isbn = csv['ISBN']
        author = csv['Author']
        publisher = csv['Publisher']
        number = csv['Number']
        books_all_details = pd.DataFrame({
            'Name': names,
            'ISBN': isbn,
            'Author': author,
            'Publisher': publisher,
            'Number': number
        })

        filename = 'booklist.xlsx'
        
        books_all_details.to_excel(filename)
        


        return render(request, 'studentlogin/done_uploading_excel.html', {'something':True})
    
    return render(request,'studentlogin/done_uploading_excel.html')


def download_books(request):

    file_path = './booklist.xlsx'
    if os.path.exists(file_path):
        with open(file_path, 'rb') as file:
            response = HttpResponse(file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename={os.path.basename(file_path)}'
            return response
    else:
        return HttpResponse("File not found.")
    

def student_login(request):
    try:
        book = openpyxl.load_workbook('./booklist.xlsx')
        sheet = book.active
        # cells = sheet['B2': 'E3']
        books_details_list = []
        for row in sheet.iter_rows(min_row=2, max_col=6, values_only=True):
            name = row[1]  # First column as 'name'
            isbn = row[2]
            author = row[3]
            publisher = row[4]
            number = row[5] 
            # print(name)
            # if name and any(detail):  # Only include rows with a name and at least one detail
            books_details_list.append({'name': name, 'isbn': isbn, 'author': author, 'publisher': publisher, 'number': number})


        

        # Generate HTML files for each detail
        generate_html_files(books_details_list)

        return render(request, 'studentlogin/student_login.html', {'details': books_details_list})
    except Exception as e:
        return render(request, 'studentlogin/student_login.html', {'error': str(e)})
    # return render(request,'studentlogin/student_login.html', )

